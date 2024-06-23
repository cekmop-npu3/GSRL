from asyncio import run, TaskGroup
from os import remove, PathLike
from openpyxl.worksheet.worksheet import Worksheet
from typing import Self
from functools import reduce

from ScoreFromPdf import PdfFormatter, Extract, Excel, Task


class TaskManager:
    async def __aenter__(self) -> Self:
        return self

    @staticmethod
    def prepareRows(ws, ws2):
        newList = list(ws.iter_rows(min_row=2, values_only=True) if isinstance(ws, Worksheet) else ws)
        newList.extend(list(ws2.iter_rows(min_row=2, values_only=True)))
        return newList

    def sumRows(self, c, d):
        if isinstance(c, tuple):
            return self.sumRows({c[0]: c[1]}, d)
        else:
            if d[0] in c:
                c.update({d[0]: str(float(c.get(d[0]).replace(',', '.')) + float(d[1].replace(',', '.'))).replace('.', ',')})
            else:
                c.update({d[0]: d[1]})
            return c

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        with Excel() as wb:
            if len(wb.worksheets) > 1:
                ws: Worksheet = wb.create_sheet('Сумма')
                for row in sorted(map(lambda item: (item[0], item[1]), reduce(self.sumRows, reduce(self.prepareRows, wb.worksheets)).items()), key=lambda x: float(x[1].replace(',', '.')), reverse=True):
                    ws.append(row)

    async def gatherTasks(self, tasks: list[Task]):
        async with TaskGroup() as tg:
            [tg.create_task(self.task(**task)) for task in tasks]

    @staticmethod
    async def task(file: str | bytes | PathLike, filename: str) -> Extract:
        filename = await PdfFormatter(file, filename).format()
        with Excel() as wb:
            ws: Worksheet = wb.create_sheet(filename)
            for row in (results := Extract(filename)):
                ws.append(list(row.values()))
        remove(filename)
        return results


async def main():
    async with TaskManager() as tm:
        await tm.gatherTasks([
            {'file': 'https://gsrl.by/wp-content/uploads/2024/06/protokol-matematika-2024.pdf', 'filename': 'protokol-matematika-2024.pdf'},
            {'file': 'https://gsrl.by/wp-content/uploads/2024/06/protokol_22_fran.pdf', 'filename': 'protokol_22_fran.pdf'}
        ])


if __name__ == '__main__':
    run(main())
